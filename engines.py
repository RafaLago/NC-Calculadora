# =============================================================================
# ENGINES - CALCULATIONS & PDF GENERATION
# Pure business logic. No Streamlit imports here.
# =============================================================================
import re
import unicodedata
from fpdf import FPDF

# --- SECTION: FORMATTING HELPERS ---
def format_real(valor):
    """Formats a float to Brazilian currency string: R$ 1.234,56"""
    return f"R$ {valor:,.2f}".replace(",", "x").replace(".", ",").replace("x", ".")

def format_id_or_phone(text, type_format):
    """Applies input masks. type_format: 'phone' | 'tax_id' (CPF or CNPJ)."""
    nums = re.sub(r'\D', '', text)
    if type_format == "phone":
        if len(nums) <= 10:
            return re.sub(r'(\d{2})(\d{4})(\d{4})', r'(\1) \2-\3', nums)
        return re.sub(r'(\d{2})(\d{5})(\d{4})', r'(\1) \2-\3', nums)
    elif type_format == "tax_id":
        if len(nums) <= 11:
            return re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1.\2.\3-\4', nums)
        return re.sub(r'(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})', r'\1.\2.\3/\4-\5', nums)
    return nums

def _safe_latin1(text):
    """Normalizes a string to latin-1 safe characters for fpdf1 compatibility.
    Decomposes accented chars (e.g. ã→a) so the PDF never crashes on
    Brazilian names. Remove this function once fully migrated to fpdf2."""
    return unicodedata.normalize('NFKD', str(text)).encode('latin-1', 'ignore').decode('latin-1')

# --- SECTION: CALCULATION ENGINES ---
def _calc_base(largura, altura, preco, tipo, max_dim=3000):
    """Base calculator for all area-priced products.
    Returns a result dict with formatted values and a validity flag."""
    area  = (largura / 1000) * (altura / 1000)
    valid = largura <= max_dim and altura <= max_dim
    prob  = ""
    if not valid:
        if largura > max_dim and altura > max_dim:
            prob = "Largura e Altura excedem o máximo"
        else:
            prob = "Largura excede o máximo" if largura > max_dim else "Altura excede o máximo"
    return {
        "tipo":     tipo,
        "area_m2":  area,
        "br_price": format_real(area * preco),
        "br_m2":    format_real(preco),
        "is_valid": valid,
        "max_dim":  max_dim,
        "problema": prob,
    }

def calcular_flexdoor(largura, altura):
    """Calculates Flexdoor pricing. Dupla if largura > 1200mm, else Simples."""
    preco = 1790.0 if largura > 1200 else 1590.0
    tipo  = "Dupla" if largura > 1200 else "Simples"
    return _calc_base(largura, altura, preco, tipo)

def calcular_pvc(largura, altura):
    """Calculates PVC Curtain pricing. No dimension limit."""
    res = _calc_base(largura, altura, 260.0, "Cortina PVC", max_dim=99999)
    res["espessura_especial"] = altura > 4000
    return res

def calcular_peca_pvc(largura, altura):
    """Calculates PVC Piece pricing."""
    return _calc_base(largura, altura, 790.0, "Peça PVC")

def calcular_deslocamento(km):
    """Calculates travel fee: km * 2 (round trip) * R$3/km."""
    return {"br_valor": format_real(km * 2 * 3)}

def calcular_acomodacao(dias, funcionarios):
    """Calculates accommodation fee with 10% overhead.
    Formula: ((R$350/day) + (R$100/day * staff)) * 1.1"""
    return {"br_valor": format_real(((350 * dias) + (100 * dias * funcionarios)) * 1.1)}

# --- SECTION: PDF GENERATION ---
def gerar_pdf_orcamento(dados, servicos, produtos, obs):
    """Generates a quote PDF in memory and returns raw bytes.

    Args:
        dados:    dict with keys: nome, empresa, email, fone, cnpj, dimensoes, area
        servicos: list of dicts with keys: item, preco, qtd
        produtos: list of dicts with keys: item, preco, qtd
        obs:      str, additional observations

    Returns:
        bytes: raw PDF content ready for st.download_button
    """
    pdf = FPDF()
    pdf.add_page()

    # Header
    pdf.set_font("Arial", "B", 16)
    pdf.cell(190, 10, _safe_latin1("Base de Orçamento - NC Portas"), ln=True, align='C')
    pdf.ln(10)

    # Client Info Block
    pdf.set_font("Arial", "", 10)
    client_block = (
        f"Cliente: {dados['nome']} | Empresa: {dados['empresa']}\n"
        f"Email: {dados['email']} | Fone: {dados['fone']}\n"
        f"CNPJ: {dados['cnpj']}\n"
        f"Vão Informado: {dados['dimensoes']} | Área: {dados['area']}"
    )
    pdf.multi_cell(190, 7, _safe_latin1(client_block), border=0)

    # Item Tables (Services and Products)
    for label, items in [("Serviços", servicos), ("Produtos", produtos)]:
        if not items:
            continue
        pdf.ln(8)

        # Table Section Label
        pdf.set_font("Arial", "B", 12)
        pdf.cell(190, 10, _safe_latin1(label), ln=True, align='L')

        # Table Header Row
        pdf.set_fill_color(240, 240, 240)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(140, 7, _safe_latin1(" Descrição"),   border=1, fill=True)
        pdf.cell(30,  7, "Preco Unit.",                border=1, fill=True, align='L')
        pdf.cell(20,  7, "Qtd",                        border=1, fill=True, align='C', ln=True)

        # Table Data Rows
        pdf.set_font("Arial", "", 10)
        for item in items:
            pdf.cell(140, 7, _safe_latin1(f" {item['item']}"), border=1)
            pdf.cell(30,  7, format_real(item['preco']),        border=1, align='L')
            pdf.cell(20,  7, str(item['qtd']),                  border=1, align='C', ln=True)

    # Observations Block
    if obs:
        pdf.ln(10)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(190, 10, _safe_latin1("Observações"), ln=True)
        pdf.set_font("Arial", "", 10)
        pdf.multi_cell(190, 7, _safe_latin1(obs), border=1)

    # Output to memory buffer
    pdf_out = pdf.output(dest='S')
    return bytes(pdf_out) if isinstance(pdf_out, bytearray) else pdf_out.encode('latin-1')

# --- SECTION: EXCEL REPORT GENERATION ---
# Template paths — model files must sit alongside this module in production.
# The _inject_cf helper post-processes the openpyxl ZIP output to:
#   1. Replace styles.xml and theme1.xml with the model's versions (preserves
#      all DXF entries, theme colors, fonts, and number formats exactly).
#   2. Inject conditional formatting XML into the sheet XML.
# This is necessary because openpyxl cannot write theme-referenced colors or
# DXF-based conditional formatting without corrupting the file.

import io, os, zipfile, re as _re
_BASE = os.path.dirname(__file__)

_NFE_TEMPLATE  = os.path.join(_BASE, "NC_Portas_-_NFe_-_Relatório_Março.xlsx")
_NFSE_TEMPLATE = os.path.join(_BASE, "NC_Portas_-_NFSe_-_Relatório_Março.xlsx")


def _inject_styles_and_cf(xlsx_bytes, template_path, cf_xml, last_data_row):
    """Post-processes an openpyxl-generated xlsx buffer:
    - Replaces styles.xml and theme1.xml with those from template_path.
    - Injects cf_xml into the sheet XML after </sheetData>.
    - Updates the table ref to cover rows 1..last_data_row.

    Args:
        xlsx_bytes:    bytes from openpyxl wb.save()
        template_path: path to the model .xlsx file
        cf_xml:        str — full conditionalFormatting XML block(s) to inject
        last_data_row: int — last row with data (for table ref update)

    Returns:
        bytes: patched xlsx content
    """
    # Read template assets
    with zipfile.ZipFile(template_path, 'r') as tmpl:
        styles_xml = tmpl.read('xl/styles.xml')
        theme_xml  = tmpl.read('xl/theme/theme1.xml')

    # Read generated file and patch it
    src  = io.BytesIO(xlsx_bytes)
    dest = io.BytesIO()

    with zipfile.ZipFile(src, 'r') as zin, \
         zipfile.ZipFile(dest, 'w', compression=zipfile.ZIP_DEFLATED) as zout:

        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == 'xl/styles.xml':
                data = styles_xml

            elif item.filename == 'xl/theme/theme1.xml':
                data = theme_xml

            elif item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
                text = data.decode('utf-8')
                # Inject CF after </sheetData> (before <pageMargins or end of worksheet)
                insert_point = '</sheetData>'
                if insert_point in text:
                    # Remove any CF blocks openpyxl may have written
                    text = _re.sub(r'<conditionalFormatting.*?</conditionalFormatting>', '',
                                   text, flags=_re.DOTALL)
                    text = text.replace(insert_point,
                                        insert_point + cf_xml, 1)
                data = text.encode('utf-8')

            elif 'xl/tables/' in item.filename:
                # Update table ref to match actual data range
                text = data.decode('utf-8')
                # Replace ref="A1:X<old_row>" with correct last row
                text = _re.sub(r'ref="(A1:[A-Z]+)\d+"',
                               lambda m: f'ref="{m.group(1)}{last_data_row}"', text)
                # Also fix autoFilter ref
                text = _re.sub(r'(<autoFilter ref="A1:[A-Z]+)\d+"',
                               lambda m: f'{m.group(1)}{last_data_row}"', text)
                data = text.encode('utf-8')

            zout.writestr(item, data)

    return dest.getvalue()


def gerar_relatorio_nfe(df, mes_ano):
    """Generates the NF-e NC monthly report Excel file in memory.
    Formatting, fonts, theme colors and conditional formatting are copied
    exactly from the model file.

    Args:
        df:      DataFrame filtered for the target month.
        mes_ano: str label e.g. 'Março 2026'

    Returns:
        bytes: raw .xlsx content ready for st.download_button
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "NF-e NC"

    # (header, df_col, width, h_align, d_align)
    COL_DEFS = [
        ("CLIENTE",    "CLIENTE",    52.625, "left",   "left"),
        ("NF-e",       "NF-e",       16.125, "center", "center"),
        ("VALOR NF-e", "VALOR NF-e", 19.625, "center", "center"),
        ("CFOP",       "CFOP",       10.875, "center", "center"),
        ("NATUREZA",   "NATUREZA",   58.875, "center", "center"),
        ("CHAVE",      "CHAVE",      52.5,   "center", "center"),
        ("EMISSÃO",    "EMISSÃO",    14.125, "center", "center"),
        ("STATUS",     "STATUS",     13.5,   "center", "center"),
    ]

    _thick  = Side(border_style="thick")
    _medium = Side(border_style="medium")
    _none   = Side(border_style=None)
    _col_left = {2: _thick, 4: _medium, 6: _medium, 8: _medium}

    for i, (_, _, w, _, _) in enumerate(COL_DEFS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_format.defaultRowHeight = 16.5

    # Header
    for ci, (hdr, _, _, ha, _) in enumerate(COL_DEFS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal=ha)
        cell.border    = Border(left=_col_left.get(ci, _none))

    # Data rows (fills are placeholders — replaced by template styles.xml)
    fill_w = PatternFill("solid", fgColor="FFFFFF")
    fill_g = PatternFill("solid", fgColor="F2F2F2")

    for ri, (_, row) in enumerate(df.iterrows()):
        er   = ri + 2
        fill = fill_w if ri % 2 == 0 else fill_g
        for ci, (_, col, _, _, da) in enumerate(COL_DEFS, 1):
            val = row.get(col, "")
            if col == "EMISSÃO" and hasattr(val, "strftime"):
                val = val.strftime("%d/%m/%Y")
            elif col == "EMISSÃO" and val == val and val:
                import pandas as pd
                try: val = pd.to_datetime(val).strftime("%d/%m/%Y")
                except Exception: pass
            cell = ws.cell(row=er, column=ci,
                           value=(val if val == val else ""))
            cell.alignment = Alignment(horizontal=da)
            cell.fill      = fill
            cell.border    = Border(left=_col_left.get(ci, _none))

    last_data_row = len(df) + 1

    # Table
    tbl = Table(displayName="NFe", ref=f"A1:H{last_data_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium15", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)

    # Footer (skip 1 blank row)
    fr      = last_data_row + 2
    _bold   = Font(bold=True)
    _mb     = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)
    for label, formula, r in [
        ("Quantidade Notas", f"=COUNT(NFe[NF-e])",     fr),
        ("Valor Total",      f"=SUM(NFe[VALOR NF-e])", fr + 1),
    ]:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = _bold; c1.border = _mb
        c2 = ws.cell(row=r, column=2, value=formula)
        c2.font = _bold; c2.border = _mb
        c2.alignment = Alignment(horizontal="center")

    # Save to buffer then inject template styling + CF
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    # Conditional formatting XML (sqref updated to match actual data range)
    cf_xml = (
        f'<conditionalFormatting sqref="B2:H{last_data_row}">'
        f'<cfRule type="expression" dxfId="2" priority="1">'
        f'<formula>$H2="Em Andamento"</formula></cfRule>'
        f'<cfRule type="expression" dxfId="1" priority="2">'
        f'<formula>$H2="Emitida"</formula></cfRule>'
        f'<cfRule type="expression" dxfId="0" priority="3">'
        f'<formula>$H2="Cancelada"</formula></cfRule>'
        f'</conditionalFormatting>'
    )

    return _inject_styles_and_cf(raw, _NFE_TEMPLATE, cf_xml, last_data_row)


def gerar_relatorio_nfse(df, mes_ano):
    """Generates the NFS-e NC monthly report Excel file in memory.
    Formatting, fonts, theme colors and conditional formatting are copied
    exactly from the model file.

    Args:
        df:      DataFrame filtered for the target month.
        mes_ano: str label e.g. 'Março 2026'

    Returns:
        bytes: raw .xlsx content ready for st.download_button
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "NFS-e NC"

    COL_DEFS = [
        ("CLIENTE",       "CLIENTE",       49.625, "left",   "left"),
        ("NFS-e",         "NFS-e",         15.0,   "center", "center"),
        ("VALOR NFS-e",   "VALOR NFS-e",   20.625, "center", "center"),
        ("DPS",           "DPS",           9.125,  "center", "center"),
        ("CHAVE",         "CHAVE",         58.875, "center", "center"),
        ("DATA",          "DATA",          11.375, "center", "center"),
        ("STATUS",        "STATUS",        13.25,  "center", "center"),
        ("OBSERVAÇÕES",   "OBSERVAÇÕES",   19.125, "left",   "left"),
        ("RETER ISS",     "RETER ISS",     15.25,  "center", "center"),
        ("ISS",           "ISS",           12.125, "center", "center"),
        ("RETER INSS",    "RETER INSS",    18.875, "center", "center"),
        ("INSS",          "INSS",          13.875, "center", "center"),
        ("VALOR LIQUIDO", "VALOR LIQUIDO", 23.625, "center", "center"),
    ]

    _thick  = Side(border_style="thick")
    _medium = Side(border_style="medium")
    _none   = Side(border_style=None)
    _col_left = {3: _medium, 5: _medium, 7: _medium, 9: _thick}

    for i, (_, _, w, _, _) in enumerate(COL_DEFS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_format.defaultRowHeight = 16.5

    # Header
    for ci, (hdr, _, _, ha, _) in enumerate(COL_DEFS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal=ha)
        cell.border    = Border(left=_col_left.get(ci, _none))

    fill_w = PatternFill("solid", fgColor="FFFFFF")
    fill_g = PatternFill("solid", fgColor="F2F2F2")

    for ri, (_, row) in enumerate(df.iterrows()):
        er   = ri + 2
        fill = fill_w if ri % 2 == 0 else fill_g
        for ci, (_, col, _, _, da) in enumerate(COL_DEFS, 1):
            val = row.get(col, "")
            if col == "DATA" and hasattr(val, "strftime"):
                val = val.strftime("%d/%m/%Y")
            elif col == "DATA" and val == val and val:
                import pandas as pd
                try: val = pd.to_datetime(val).strftime("%d/%m/%Y")
                except Exception: pass
            cell = ws.cell(row=er, column=ci,
                           value=(val if val == val else ""))
            cell.alignment = Alignment(horizontal=da)
            cell.fill      = fill
            cell.border    = Border(left=_col_left.get(ci, _none))

    last_data_row = len(df) + 1

    tbl = Table(displayName="NFSe", ref=f"A1:M{last_data_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium15", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)

    fr    = last_data_row + 2
    _bold = Font(bold=True)
    _mb   = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)
    for label, formula, r in [
        ("Quantidade Notas", f"=COUNT(NFSe[NFS-e])",         fr),
        ("Valor Total",      f"=SUM(NFSe[VALOR NFS-e])",     fr + 1),
        ("Valor Liquido",    f"=SUM(NFSe[VALOR LIQUIDO])",   fr + 2),
    ]:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = _bold; c1.border = _mb
        c2 = ws.cell(row=r, column=2, value=formula)
        c2.font = _bold; c2.border = _mb
        c2.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    # 4 CF blocks with dynamic sqref
    cf_xml = (
        # STATUS colors on columns B-H
        f'<conditionalFormatting sqref="B2:H{last_data_row}">'
        f'<cfRule type="expression" dxfId="9" priority="8">'
        f'<formula>$G2="Cancelada"</formula></cfRule>'
        f'<cfRule type="expression" dxfId="8" priority="9">'
        f'<formula>$G2="Substituida"</formula></cfRule>'
        f'<cfRule type="expression" dxfId="7" priority="10">'
        f'<formula>$G2="Emitida"</formula></cfRule>'
        f'</conditionalFormatting>'
        # RETER ISS colors on columns I-J
        f'<conditionalFormatting sqref="I2:J{last_data_row}">'
        f'<cfRule type="expression" dxfId="6" priority="6">'
        f'<formula>$I2="NÃO"</formula></cfRule>'
        f'<cfRule type="expression" dxfId="5" priority="7">'
        f'<formula>$I2="SIM"</formula></cfRule>'
        f'</conditionalFormatting>'
        # RETER INSS colors on columns K-L
        f'<conditionalFormatting sqref="K2:L{last_data_row}">'
        f'<cfRule type="expression" dxfId="4" priority="4">'
        f'<formula>$K2="NÃO"</formula></cfRule>'
        f'<cfRule type="expression" dxfId="3" priority="5">'
        f'<formula>$K2="SIM"</formula></cfRule>'
        f'</conditionalFormatting>'
        # VALOR LIQUIDO color on column M
        f'<conditionalFormatting sqref="M2:M{last_data_row}">'
        f'<cfRule type="expression" dxfId="2" priority="1">'
        f'<formula>$G2="Cancelada"</formula></cfRule>'
        f'<cfRule type="expression" dxfId="1" priority="2">'
        f'<formula>$G2="Substituida"</formula></cfRule>'
        f'<cfRule type="expression" dxfId="0" priority="3">'
        f'<formula>$G2="Emitida"</formula></cfRule>'
        f'</conditionalFormatting>'
    )

    return _inject_styles_and_cf(raw, _NFSE_TEMPLATE, cf_xml, last_data_row)
